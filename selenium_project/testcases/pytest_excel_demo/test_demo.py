import pytest
import openpyxl

#with open('.\testcases\pytest_excel_demo\test_demo.xlsx', 'rb') as f:
wb = openpyxl.load_workbook(r'.\testcases\pytest_excel_demo\test_demo.xlsx')
sheet_1=wb['名单表']


'''name_1=sheet_1['A2'].value
name_2=sheet_1['A3'].value
name_3=sheet_1['A4'].value'''

'''name_1=sheet_1.cell(row=2, column=1).value
name_2=sheet_1.cell(row=3, column=1).value
name_3=sheet_1.cell(row=4, column=1).value

print(name_1,name_2,name_3)'''

list_data_0 = []
for column in sheet_1.iter_cols(min_col=1, max_col=2, min_row=2, max_row=4, values_only=True):
    list_data_0.append(column)

list_data1=[]
def get_data(sheet_1):
    for row in sheet_1.iter_rows(min_row=2, max_row=4, min_col=1, max_col=2, values_only=True):
        yield row[0],row[1]

#print(list_data1)


@pytest.mark.parametrize('name', list_data_0[0])
def test_name(name):
    print(name)


@pytest.mark.parametrize('user,age', get_data(sheet_1))
def test_user_info(user,age):
    user_name=user
    user_age=age
    print(user_name,user_age)


if __name__ == '__main__':
    pytest.main(['-sv', r'.\testcases\pytest_excel_demo\test_demo.py'])






